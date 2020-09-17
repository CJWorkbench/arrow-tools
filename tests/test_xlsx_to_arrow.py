import datetime
from pathlib import Path
import subprocess
import tempfile
from typing import Tuple, Union
import openpyxl as xl
import pyarrow
from .util import assert_table_equals


def do_convert(
    xlsx_path: Path,
    *,
    max_rows: int = 99999,
    max_columns: int = 99998,
    max_bytes_per_value: int = 99997,
    max_bytes_total: int = 999999999,
    header_rows: str = "0-1",
    header_rows_file: str = "",
    include_stdout: bool = False
) -> Union[pyarrow.Table, Tuple[pyarrow.Table, bytes]]:
    with tempfile.NamedTemporaryFile(suffix=".arrow") as arrow_file:
        args = [
            "/usr/bin/xlsx-to-arrow",
            "--max-rows",
            str(max_rows),
            "--max-columns",
            str(max_columns),
            "--max-bytes-per-value",
            str(max_bytes_per_value),
            "--max-bytes-total",
            str(max_bytes_total),
            "--header-rows",
            header_rows,
            "--header-rows-file",
            header_rows_file,
            xlsx_path.as_posix(),
            Path(arrow_file.name).as_posix(),
        ]
        try:
            result = subprocess.run(
                args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True
            )
        except subprocess.CalledProcessError as err:
            # Rewrite error so it's easy to read in test-result stack trace
            raise RuntimeError(
                "Process failed with code %d: %s"
                % (
                    err.returncode,
                    (
                        err.stdout.decode("utf-8", errors="replace")
                        + err.stderr.decode("utf-8", errors="replace")
                    ),
                )
            ) from None

        assert result.stderr == b""
        with pyarrow.ipc.open_file(arrow_file.name) as result_reader:
            table = result_reader.read_all()
        if include_stdout:
            return table, result.stdout
        else:
            assert result.stdout == b""
            return table


def do_convert_data(
    workbook: xl.Workbook, **kwargs
) -> Union[pyarrow.Table, Union[pyarrow.Table, str]]:
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as xlsx_file:
        workbook.save(filename=xlsx_file.name)
        return do_convert(Path(xlsx_file.name), **kwargs)


# This is hard to test, since it's really an invalid Excel file
# def test_no_sheets_is_error():
#     # https://openpyxl.readthedocs.io/en/stable/optimized.html#write-only-mode
#     # ... to create a workbook with no worksheets
#     workbook = xl.Workbook()
#     workbook.remove(workbook.active)
#     workbook.get_active_sheet = lambda: None
#     result, stdout = do_convert_data(workbook, include_stdout=True)
#     assert_table_equals(result, pyarrow.table({}))
#     assert stdout == b"Excel file has no worksheets\n"


def test_empty_sheet():
    workbook = xl.Workbook()
    assert_table_equals(do_convert_data(workbook, header_rows=""), pyarrow.table({}))


def test_empty_sheet_no_header_row():
    workbook = xl.Workbook()
    assert_table_equals(do_convert_data(workbook, header_rows="0-1"), pyarrow.table({}))


def test_number_columns():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append([1, 1.1])
    sheet.append([2, 2.2])
    sheet.append([3, 3.3])
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": [1.0, 2.0, 3.0], "B": [1.1, 2.2, 3.3]}),
    )


# openpyxl doesn't write shared strings
# def test_shared_string_column():
#     workbook = xl.Workbook()
#     sheet = workbook.active
#     sheet.append(["a"])
#     sheet.append(["b"])
#     assert_table_equals(
#         do_convert_data(workbook, header_rows=""),
#         pyarrow.table({"A": ["a", "b"]})
#     )


def test_inline_str_column():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append(["a"])
    sheet.append(["b"])
    assert_table_equals(
        do_convert_data(workbook, header_rows=""), pyarrow.table({"A": ["a", "b"]})
    )


def test_date_and_datetime_columns():
    workbook = xl.Workbook()
    sheet = workbook.active
    # These dates are chosen specially -- double precision can't represent
    # every arbitrary number of microseconds accurately (let alone
    # nanoseconds), but the math happens to work for these datetimes.
    #
    # (We aren't testing rounding here.)
    sheet.append(
        [datetime.date(2020, 1, 25), datetime.datetime(2020, 1, 25, 17, 25, 30, 128000)]
    )
    sheet.append(
        [datetime.date(2020, 1, 26), datetime.datetime(2020, 1, 25, 17, 26, 27, 256)]
    )
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table(
            {
                "A": pyarrow.array(
                    [datetime.datetime(2020, 1, 25), datetime.datetime(2020, 1, 26)],
                    pyarrow.timestamp("ns"),
                ),
                "B": pyarrow.array(
                    [
                        datetime.datetime(2020, 1, 25, 17, 25, 30, 128000),
                        datetime.datetime(2020, 1, 25, 17, 26, 27, 256),
                    ],
                    pyarrow.timestamp("ns"),
                ),
            }
        ),
    )


def test_datetime_overflow():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append([datetime.date(1100, 1, 1), datetime.date(1901, 1, 1)])
    sheet.append([datetime.date(1901, 1, 1), datetime.date(3000, 1, 1)])
    result, stdout = do_convert_data(workbook, include_stdout=True, header_rows="")
    assert_table_equals(
        result,
        pyarrow.table(
            {
                "A": pyarrow.array(
                    [None, datetime.datetime(1901, 1, 1)], pyarrow.timestamp("ns")
                ),
                "B": pyarrow.array(
                    [datetime.datetime(1901, 1, 1), None], pyarrow.timestamp("ns")
                ),
            }
        ),
    )
    assert (
        stdout
        == b"replaced out-of-range with null for 2 Timestamps; see row 0 column A\n"
    )


def test_convert_datetime_to_string_and_report():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = datetime.date(1981, 1, 1)
    sheet["A2"] = "hi"
    result, stdout = do_convert_data(workbook, include_stdout=True, header_rows="")
    assert_table_equals(
        result,
        pyarrow.table({"A": ["1981-01-01", "hi"]}),
    )
    assert stdout == b"interpreted 1 Timestamps as String; see row 0 column A\n"


def test_datetime_do_not_convert_to_string_when_value_is_whitespace():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = datetime.date(1981, 1, 1)
    sheet["A2"] = "  "
    sheet["A3"] = datetime.date(1983, 1, 1)
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table(
            {
                "A": pyarrow.array(
                    [
                        datetime.datetime(1981, 1, 1),
                        None,
                        datetime.datetime(1983, 1, 1),
                    ],
                    pyarrow.timestamp("ns"),
                ),
            }
        ),
    )


def test_skip_null_values():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A2"] = 3.0
    sheet["A4"] = 4.0
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": [None, 3.0, None, 4.0]}),
    )


def test_skip_null_columns():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 3.0
    sheet["A2"] = 3.0
    sheet["D1"] = 4.0
    sheet["D2"] = 4.0
    result, stdout = do_convert_data(workbook, header_rows="", include_stdout=True)
    assert_table_equals(
        result,
        pyarrow.table(
            {
                "A": [3.0, 3.0],
                "B": pyarrow.array([None, None], pyarrow.utf8()),
                "C": pyarrow.array([None, None], pyarrow.utf8()),
                "D": [4.0, 4.0],
            }
        ),
    )
    assert stdout == b"chose string type for null column B and more\n"


def test_backfill_column_at_end():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = 3.0
    sheet["B1"] = 4.0
    sheet["B2"] = 4.0
    sheet["B3"] = 4.0
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": [3.0, None, None], "B": [4.0, 4.0, 4.0]}),
    )


def test_bool_becomes_str():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append([True])
    sheet.append([False])
    result, stdout = do_convert_data(workbook, header_rows="", include_stdout=True)
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": ["TRUE", "FALSE"]}),
    )


def test_invalid_zipfile():
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tf:
        path = Path(tf.name)
        path.write_bytes(b"12345")
        result, stdout = do_convert(path, include_stdout=True)
    assert_table_equals(result, pyarrow.table({}))
    assert stdout == b"Invalid XLSX file: xlnt::exception : failed to find zip header\n"


def test_xml_error():
    result, stdout = do_convert(
        Path(__file__).parent / "files" / "xml-required-attribute-missing-in-rels.xlsx",
        include_stdout=True,
    )
    assert_table_equals(result, pyarrow.table({}))
    assert (
        stdout
        == b"Invalid XLSX file: xl/_rels/workbook.xml.rels:2:84: error: attribute 'Target' expected\n"
    )


def test_skip_rows():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Hi"
    sheet["A2"] = 1.0
    sheet["A4"] = 1.0
    result, stdout = do_convert_data(
        workbook, max_rows=1, header_rows="0-1", include_stdout=True
    )
    assert_table_equals(result, pyarrow.table({"A": [1.0]}))
    assert stdout == b"skipped 2 rows (after row limit of 1)\n"


def test_skip_columns():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "a"
    sheet["B1"] = "b"
    sheet["C1"] = "c"
    result, stdout = do_convert_data(
        workbook, max_columns=1, header_rows="", include_stdout=True
    )
    assert_table_equals(result, pyarrow.table({"A": ["a"]}))
    assert stdout == b"skipped column B and more (after column limit of 1)\n"


def test_header_rows_convert_to_str():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append([datetime.date(2020, 1, 25), 123.4213, 123.4213, None, ""])
    sheet["A1"].number_format = "dd-mmm-yyyy"
    sheet["C1"].number_format = "#.00"
    sheet.append(["a", "b", "c", "d", "e"])
    with tempfile.NamedTemporaryFile(suffix="-headers.arrow") as header_file:
        # ignore result
        do_convert_data(workbook, header_rows="0-1", header_rows_file=header_file.name)
        with pyarrow.ipc.open_file(header_file.name) as header_reader:
            header_table = header_reader.read_all()
    assert_table_equals(
        header_table,
        pyarrow.table(
            {
                "A": ["25-Jan-2020"],
                "B": ["123.4213"],
                "C": ["123.42"],
                "D": pyarrow.array([None], pyarrow.utf8()),
                "E": [""],
            }
        ),
    )


def test_header_rows():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append(["ColA", "ColB"])
    sheet.append(["a", "b"])
    with tempfile.NamedTemporaryFile(suffix="-headers.arrow") as header_file:
        result = do_convert_data(
            workbook, header_rows="0-1", header_rows_file=header_file.name
        )
        with pyarrow.ipc.open_file(header_file.name) as header_reader:
            header_table = header_reader.read_all()
    assert_table_equals(result, pyarrow.table({"A": ["a"], "B": ["b"]}))
    assert_table_equals(header_table, pyarrow.table({"A": ["ColA"], "B": ["ColB"]}))


def test_header_truncated():
    workbook = xl.Workbook()
    sheet = workbook.active
    sheet.append(["xy1", "xy2"])
    sheet.append(["a", "b"])
    with tempfile.NamedTemporaryFile(suffix="-headers.arrow") as header_file:
        result, stdout = do_convert_data(
            workbook,
            max_bytes_per_value=2,
            header_rows="0-1",
            header_rows_file=header_file.name,
            include_stdout=True,
        )
        with pyarrow.ipc.open_file(header_file.name) as header_reader:
            header_table = header_reader.read_all()
    assert_table_equals(result, pyarrow.table({"A": ["a"], "B": ["b"]}))
    assert_table_equals(header_table, pyarrow.table({"A": ["xy"], "B": ["xy"]}))
    assert stdout == b"".join(
        [b"truncated 2 values (value byte limit is 2; see row 0 column A)\n"]
    )


def test_values_truncated():
    workbook = xl.Workbook()
    workbook.active.append(["abcde", "fghijklmn", "opq"])
    result, stdout = do_convert_data(
        workbook,
        max_bytes_per_value=3,
        header_rows="",
        include_stdout=True,
    )
    assert_table_equals(
        result, pyarrow.table({"A": ["abc"], "B": ["fgh"], "C": ["opq"]})
    )
    assert stdout == b"truncated 2 values (value byte limit is 3; see row 0 column A)\n"


def test_truncate_do_not_cause_invalid_utf8():
    workbook = xl.Workbook()
    for s in [
        # Examples from https://en.wikipedia.org/wiki/UTF-8
        "AAAA",
        "AA\u00A2",  # ¢ (2 bytes) -- keep
        "AAA\u00A2",  # ¢ (2 bytes) -- drop both bytes
        "A\u0939",  # ह (3 bytes) -- keep
        "AA\u0939",  # ह (3 bytes) -- drop all three bytes
        "AAA\u0939",  # ह (3 bytes) -- drop all three bytes
        "\U00010348",  # 𐍈 (4 bytes) -- keep
        "A\U00010348",  # 𐍈 (4 bytes) -- drop all four bytes
        "AA\U00010348",  # 𐍈 (4 bytes) -- drop all four bytes
        "AAA\U00010348",  # 𐍈 (4 bytes) -- drop all four bytes
    ]:
        workbook.active.append([s])

    result, stdout = do_convert_data(
        workbook,
        max_bytes_per_value=4,
        header_rows="",
        include_stdout=True,
    )
    expected = pyarrow.table(
        {
            "A": [
                "AAAA",
                "AA\u00A2",
                "AAA",
                "A\u0939",
                "AA",
                "AAA",
                "\U00010348",
                "A",
                "AA",
                "AAA",
            ]
        }
    )
    assert_table_equals(result, expected)
    assert stdout == b"truncated 6 values (value byte limit is 4; see row 2 column A)\n"


def test_convert_float_to_string_and_report():
    workbook = xl.Workbook()
    workbook.active["A1"] = 3.4
    workbook.active["A2"] = "s"
    workbook.active["A3"] = -2.2
    result, stdout = do_convert_data(
        workbook,
        header_rows="",
        include_stdout=True,
    )
    assert_table_equals(result, pyarrow.table({"A": ["3.4", "s", "-2.2"]}))
    assert stdout == b"interpreted 2 Numbers as String; see row 0 column A\n"


def test_float_do_not_convert_to_string_when_value_is_whitespace():
    workbook = xl.Workbook()
    workbook.active["A1"] = 3.4
    workbook.active["A2"] = "  "
    workbook.active["A3"] = -2.2
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": [3.4, None, -2.2]}),
    )


def test_float_do_not_convert_to_string_when_first_cell_is_whitespace():
    workbook = xl.Workbook()
    workbook.active["A1"] = "  "
    workbook.active["A2"] = -2.2
    assert_table_equals(
        do_convert_data(workbook, header_rows=""),
        pyarrow.table({"A": [None, -2.2]}),
    )


def test_float_convert_to_string_preserve_previously_ignored_whitespace():
    workbook = xl.Workbook()
    workbook.active["A1"] = 3.4
    workbook.active["A2"] = "  "
    workbook.active["A3"] = "x"
    result, stdout = do_convert_data(
        workbook,
        header_rows="",
        include_stdout=True,
    )
    assert_table_equals(result, pyarrow.table({"A": ["3.4", "  ", "x"]}))
    # Whitespace isn't "counted" as Number, even though we treated it as a null Number
    assert stdout == b"interpreted 1 Numbers as String; see row 0 column A\n"


def test_stop_after_byte_total_limit():
    workbook = xl.Workbook()
    workbook.active.append(["abcd", "efgh"])
    workbook.active.append(["ijkl", "mnop"])
    result, stdout = do_convert_data(
        workbook,
        max_bytes_total=8,
        header_rows="",
        include_stdout=True,
    )
    assert_table_equals(result, pyarrow.table({"A": ["abcd"], "B": ["efgh"]}))
    assert stdout == b"stopped at limit of 8 bytes of data\n"
